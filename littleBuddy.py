from openpyxl import Workbook
from openpyxl import load_workbook

#-----Start Config----------------#

#list of values to search for in order (will first search for first then second and so on).
# more can be added if you want 
KEYS=["First","Second","Third"]

#row to start inserting rows that are marked with keys 
ITERATOR_START=5

#row to start inserting rows that are unmarked
UNDEF_START=50

#name of the input file 
#must be in the same directory as the .py file
INPUT_FILE='sample_input1.xlsx'

#name of the output file to be created or modified
#if a file of this name does not exist then a new one will be created in the same directory as the .py file
OUTPUT_FILE='sample_output1.xlsx'

#column that contains the key values
KEY_COL=0


#-----End Config------------------#


try:

    if ITERATOR_START>= UNDEF_START:
        raise Exception('iterator is greater than undef. This will result in data being overwritten in the output file')
    iterator=ITERATOR_START
    undef=UNDEF_START
    wb = load_workbook(filename = INPUT_FILE)
    sheets = wb.sheetnames

    newWb = Workbook() 
    newSheet= newWb.sheetnames
    newWs=newWb[newSheet[0]]


    ws=wb[sheets[0]]
    for i in KEYS: 
        for row in ws.iter_rows():
        
        
            if row[KEY_COL].value==i:
                for cell in row:
                    newWs.cell(row=iterator, column=cell.column).value = cell.value
                iterator+=1
                if iterator>=UNDEF_START:
                     raise Exception('Iterator has reached range reserved for undefined rows. Iterator: '+str(iterator)+' Undef_Start: '+str(UNDEF_START))
        
           
           
          
    for row in ws.iter_rows():
        if (row[KEY_COL].value not in KEYS ):
            for cell in row:
                newWs.cell(row=undef, column=cell.column).value = cell.value
            undef+=1  

    newWb.save(filename=OUTPUT_FILE)
    
except FileNotFoundError as e:
        print("ERROR: Unable to open file. File may be outside of application directory") #Does not exist OR no read permissions
except PermissionError as e:
    print("ERROR: Output file is open. Please close it and try again") 
except Exception as e:
    print('ERROR: ', e)
            